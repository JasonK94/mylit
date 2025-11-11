#!/usr/bin/env python3
"""
GeomX Data Processor v2 - Enhanced Version with Better Error Handling
개선사항:
1. 손상된 xlsx 파일 복구 및 읽기
2. 의미없는 고유 열 필터링
3. column_mapping_template 생성 오류 수정
"""

import pandas as pd
import numpy as np
import os
import sys
import argparse
import hashlib
import json
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional
import warnings
warnings.filterwarnings('ignore')

# 공통 columns 정의 (다양한 데이터 타입별로)
COMMON_COLUMNS = {
    'WTA': [
        'SlideName', 'ScanLabel', 'ROILabel', 'SegmentLabel', 'SegmentDisplayName',
        'Origin Instrument ID', 'QCFlags', 'AOISurfaceArea', 'AOINucleiCount',
        'ROICoordinateX', 'ROICoordinateY', 'RawReads', 'AlignedReads',
        'DeduplicatedReads', 'TrimmedReads', 'StitchedReads', 'SequencingSaturation',
        'SequencingSetID', 'UMIQ30', 'RTSQ30', 'GeoMxNgsPipelineVersion',
        'LOT_Human_NGS_Whole_Transcriptome_Atlas_RNA_1_0', 'ROIID', 'SegmentID',
        'ScanWidth', 'ScanHeight', 'ScanOffsetX', 'ScanOffsetY'
    ],
    'CTA': [
        'SlideName', 'ScanLabel', 'ROILabel', 'SegmentLabel', 'SegmentDisplayName',
        'Origin Instrument ID', 'QCFlags', 'AOISurfaceArea', 'AOINucleiCount',
        'ROICoordinateX', 'ROICoordinateY', 'RawReads', 'AlignedReads',
        'DeduplicatedReads', 'TrimmedReads', 'StitchedReads', 'SequencingSaturation',
        'SequencingSetID', 'UMIQ30', 'RTSQ30', 'GeoMxNgsPipelineVersion',
        'LOT_Cancer_Transcriptome_Atlas', 'ROIID', 'SegmentID',
        'ScanWidth', 'ScanHeight', 'ScanOffsetX', 'ScanOffsetY'
    ],
    'MOUSE': [
        'SlideName', 'ScanLabel', 'ROILabel', 'SegmentLabel', 'SegmentDisplayName',
        'Origin Instrument ID', 'QCFlags', 'AOISurfaceArea', 'AOINucleiCount',
        'ROICoordinateX', 'ROICoordinateY', 'RawReads', 'AlignedReads',
        'DeduplicatedReads', 'TrimmedReads', 'StitchedReads', 'SequencingSaturation',
        'SequencingSetID', 'UMIQ30', 'RTSQ30', 'GeoMxNgsPipelineVersion',
        'LOT_Mouse_NGS_Whole_Transcriptome_Atlas_RNA_1_0', 'ROIID', 'SegmentID',
        'ScanWidth', 'ScanHeight', 'ScanOffsetX', 'ScanOffsetY'
    ]
}

# 무시할 column 패턴들 (의미없는 고유 열)
IGNORE_COLUMN_PATTERNS = [
    r'LOT_.*_RNA_\d+$',  # LOT_Human_NGS_Whole_Transcriptome_Atlas_RNA_388 등
    r'Scan_ID_\d+$',     # Scan_ID_435 등
    r'ROI_ID_\d+$',      # ROI_ID_76 등
    r'^\d+$',            # 숫자만 있는 column (Excel 오류로 생성되는 경우)
]

class GeomXProcessor:
    def __init__(self, verbose=True):
        self.verbose = verbose
        self.files_data = []
        self.unique_columns_analysis = []
        self.column_mappings = {}
        self.problematic_files = []
        
    def log(self, message, level='INFO'):
        if self.verbose:
            print(f"[{level}] {message}")
    
    def should_ignore_column(self, col_name):
        """의미없는 column인지 확인"""
        if not col_name or pd.isna(col_name):
            return True
            
        col_str = str(col_name)
        
        # 패턴 매칭으로 무시할 column 확인
        for pattern in IGNORE_COLUMN_PATTERNS:
            if re.match(pattern, col_str):
                return True
        
        return False
    
    def is_meaningless_column(self, series):
        """
        Column이 의미없는지 확인
        - 모든 값이 동일하거나
        - 모든 값이 고유하거나 (ID처럼)
        - 거의 모든 값이 NaN
        """
        non_null = series.dropna()
        
        if len(non_null) == 0:
            return True
            
        # 모든 값이 동일한 경우
        if len(non_null.unique()) == 1:
            return True
            
        # 모든 값이 고유한 경우 (90% 이상)
        if len(non_null.unique()) > len(non_null) * 0.9:
            return True
            
        # 대부분이 NaN인 경우 (90% 이상)
        if len(non_null) < len(series) * 0.1:
            return True
            
        return False
    
    def detect_data_type(self, df_columns):
        """데이터 타입 자동 감지 (WTA, CTA, MOUSE)"""
        columns_str = ' '.join(str(c) for c in df_columns if c)
        
        if 'LOT_Cancer_Transcriptome_Atlas' in columns_str:
            return 'CTA'
        elif 'LOT_Mouse_NGS' in columns_str:
            return 'MOUSE'
        else:
            return 'WTA'
    
    def extract_segment_properties(self, file_path):
        """xlsx 파일에서 SegmentProperties 시트 추출 (오류 복구 포함)"""
        try:
            # 여러 엔진 시도
            engines = ['openpyxl', 'xlrd', None]
            df = None
            successful_engine = None
            
            for engine in engines:
                try:
                    if engine:
                        excel_file = pd.ExcelFile(file_path, engine=engine)
                    else:
                        excel_file = pd.ExcelFile(file_path)
                    
                    # SegmentProperties 시트 찾기
                    sheet_name = None
                    for sheet in excel_file.sheet_names:
                        if 'SegmentProperties' in sheet:
                            sheet_name = sheet
                            break
                    
                    if not sheet_name:
                        continue
                    
                    # 데이터 읽기 시도
                    df = pd.read_excel(excel_file, sheet_name=sheet_name)
                    
                    # 데이터 검증
                    if df is not None and len(df) > 0:
                        successful_engine = engine if engine else 'default'
                        break
                        
                except Exception as e:
                    continue
            
            if df is None or df.empty:
                # 마지막 시도: 직접 읽기
                try:
                    import openpyxl
                    from openpyxl import load_workbook
                    
                    # data_only=True로 수식 대신 값을 읽음
                    wb = load_workbook(file_path, data_only=True, read_only=True)
                    
                    # SegmentProperties 시트 찾기
                    sheet_name = None
                    for name in wb.sheetnames:
                        if 'SegmentProperties' in name:
                            sheet_name = name
                            break
                    
                    if sheet_name:
                        ws = wb[sheet_name]
                        
                        # 데이터를 리스트로 변환
                        data = []
                        for row in ws.iter_rows(values_only=True):
                            # None이 아닌 값이 있는 행만 추가
                            if any(cell is not None for cell in row):
                                data.append(row)
                        
                        if data:
                            # DataFrame 생성
                            df = pd.DataFrame(data[1:], columns=data[0])
                            successful_engine = 'openpyxl_direct'
                    
                    wb.close()
                    
                except Exception as e:
                    self.log(f"Direct reading failed for {file_path}: {str(e)}", 'ERROR')
            
            if df is not None and not df.empty:
                # NaN column 제거
                df = df.loc[:, df.columns.notna()]
                
                # 숫자로만 된 column 이름 정리 (Excel 오류로 생성된 경우)
                new_columns = []
                for col in df.columns:
                    if isinstance(col, (int, float)):
                        new_columns.append(f"Column_{int(col)}")
                    else:
                        new_columns.append(str(col))
                df.columns = new_columns
                
                self.log(f"Successfully read {file_path} using {successful_engine}")
                return df
            else:
                self.log(f"Warning: Could not read SegmentProperties from {file_path}", 'WARNING')
                self.problematic_files.append(file_path)
                return None
                
        except Exception as e:
            self.log(f"Error reading {file_path}: {str(e)}", 'ERROR')
            self.problematic_files.append(file_path)
            return None
    
    def analyze_unique_columns(self, file_path):
        """파일의 고유 column 분석"""
        df = self.extract_segment_properties(file_path)
        if df is None:
            return
        
        file_name = Path(file_path).stem
        data_type = self.detect_data_type(df.columns.tolist())
        common_cols = COMMON_COLUMNS.get(data_type, COMMON_COLUMNS['WTA'])
        
        # 고유 columns 찾기 (무의미한 column 제외)
        unique_cols = []
        ignored_cols = []
        
        for col in df.columns:
            if col and col not in common_cols:
                # 무시할 패턴인지 확인
                if self.should_ignore_column(col):
                    ignored_cols.append(col)
                    continue
                
                # 의미없는 column인지 확인
                if self.is_meaningless_column(df[col]):
                    ignored_cols.append(col)
                    continue
                    
                unique_cols.append(col)
        
        self.log(f"File: {file_name}")
        self.log(f"  Data Type: {data_type}")
        self.log(f"  Total Columns: {len(df.columns)}")
        self.log(f"  Unique Columns: {len(unique_cols)}")
        self.log(f"  Ignored Columns: {len(ignored_cols)}")
        
        if ignored_cols and self.verbose:
            self.log(f"  Ignored: {ignored_cols[:5]}{'...' if len(ignored_cols) > 5 else ''}")
        
        # 각 고유 column 분석
        for col in unique_cols:
            try:
                # 데이터 타입 추론
                dtype = self.infer_column_type(df[col])
                
                # 고유값 샘플
                unique_values = df[col].dropna().unique()
                value_sample = list(unique_values[:10])  # 최대 10개
                
                self.unique_columns_analysis.append({
                    'dataset': file_name,
                    'data_type': data_type,
                    'column_name': col,
                    'dtype': dtype,
                    'unique_values': value_sample,
                    'unique_count': len(unique_values),
                    'null_count': df[col].isna().sum(),
                    'total_count': len(df)
                })
            except Exception as e:
                self.log(f"  Error analyzing column {col}: {str(e)}", 'WARNING')
        
        # 전체 데이터 저장
        self.files_data.append({
            'file_name': file_name,
            'data_type': data_type,
            'dataframe': df,
            'unique_columns': unique_cols
        })
    
    def infer_column_type(self, series):
        """Column의 데이터 타입 추론"""
        non_null = series.dropna()
        if len(non_null) == 0:
            return 'Empty'
        
        # Boolean 체크
        unique_lower = set(str(v).lower() for v in non_null.unique())
        bool_values = {'true', 'false', 'yes', 'no', '1', '0', 't', 'f', 'y', 'n'}
        if unique_lower.issubset(bool_values):
            return 'Boolean'
        
        # Numeric 체크
        try:
            pd.to_numeric(non_null)
            return 'Numeric'
        except:
            pass
        
        # Category 체크 (고유값이 전체의 10% 미만)
        if len(non_null.unique()) < len(non_null) * 0.1:
            return 'Category'
        
        return 'String'
    
    def process_directory(self, directory_path):
        """디렉토리 내 모든 xlsx 파일 처리"""
        xlsx_files = list(Path(directory_path).glob("*.xlsx"))
        self.log(f"Found {len(xlsx_files)} xlsx files")
        
        for file_path in xlsx_files:
            self.log(f"\nProcessing: {file_path.name}")
            self.analyze_unique_columns(str(file_path))
        
        if self.problematic_files:
            self.log(f"\n{len(self.problematic_files)} files had issues:", 'WARNING')
            for f in self.problematic_files:
                self.log(f"  - {Path(f).name}", 'WARNING')
    
    def save_analysis_results(self, output_dir):
        """분석 결과 저장"""
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        # 1. 고유 column 분석 결과
        if self.unique_columns_analysis:
            df_analysis = pd.DataFrame(self.unique_columns_analysis)
            
            # unique_values를 문자열로 변환
            df_analysis['unique_values'] = df_analysis['unique_values'].apply(
                lambda x: ', '.join(map(str, x)) if isinstance(x, list) else str(x)
            )
            
            output_file = output_path / 'unique_columns_analysis.csv'
            df_analysis.to_csv(output_file, index=False)
            self.log(f"\nSaved unique columns analysis to {output_file}")
            
            # 요약 통계
            print("\n=== Unique Columns Summary ===")
            summary = df_analysis.groupby(['data_type', 'column_name']).size()
            print(summary.to_string())
            
            # 문제 파일 목록 저장
            if self.problematic_files:
                problem_file = output_path / 'problematic_files.txt'
                with open(problem_file, 'w') as f:
                    f.write("Files with reading issues:\n")
                    for pf in self.problematic_files:
                        f.write(f"{pf}\n")
                self.log(f"Saved problematic files list to {problem_file}")
        else:
            self.log("No unique columns found to analyze", 'WARNING')
    
    def generate_mapping_template(self, output_dir):
        """Column 매핑 템플릿 생성 (오류 수정)"""
        if not self.unique_columns_analysis:
            self.log("No analysis data available", 'WARNING')
            return None
        
        df = pd.DataFrame(self.unique_columns_analysis)
        
        # column별 그룹화
        mapping_template = []
        for col_name in df['column_name'].unique():
            col_data = df[df['column_name'] == col_name]
            
            # 모든 고유값 수집
            all_values = set()
            for _, row in col_data.iterrows():
                values_str = row['unique_values']
                # pd.notna 오류 수정: 단일 값으로 변환
                if values_str is not None and str(values_str) != 'nan':
                    if isinstance(values_str, str):
                        values = values_str.split(', ')
                    else:
                        values = [str(values_str)]
                    all_values.update(values)
            
            mapping_template.append({
                'original_column': col_name,
                'unified_name': col_name,  # 기본값
                'data_types': ', '.join(col_data['data_type'].unique()),
                'datasets': ', '.join(col_data['dataset'].unique()),
                'sample_values': ', '.join(list(all_values)[:20]),
                'value_mapping': '',  # 사용자가 채울 부분
                'notes': ''
            })
        
        if mapping_template:
            df_template = pd.DataFrame(mapping_template)
            output_file = Path(output_dir) / 'column_mapping_template.csv'
            df_template.to_csv(output_file, index=False)
            self.log(f"Saved mapping template to {output_file}")
            return df_template
        else:
            self.log("No columns to create mapping template", 'WARNING')
            return None
    
    def apply_mappings_and_merge(self, mapping_file, output_dir):
        """매핑 적용 및 metadata 통합"""
        # 매핑 파일 읽기
        df_mapping = pd.read_csv(mapping_file)
        
        # 매핑 딕셔너리 생성
        column_mappings = {}
        value_mappings = {}
        
        for _, row in df_mapping.iterrows():
            original = row['original_column']
            unified = row['unified_name']
            column_mappings[original] = unified
            
            # 값 매핑 파싱
            if pd.notna(row.get('value_mapping', '')):
                mappings = str(row['value_mapping']).split(',')
                value_map = {}
                for mapping in mappings:
                    if '=' in mapping:
                        old, new = mapping.strip().split('=', 1)
                        value_map[old.strip()] = new.strip()
                if value_map:
                    value_mappings[original] = value_map
        
        # 모든 데이터프레임 통합
        merged_dfs = []
        
        for file_data in self.files_data:
            df = file_data['dataframe'].copy()
            df['_source_file'] = file_data['file_name']
            df['_data_type'] = file_data['data_type']
            
            # Column 이름 매핑
            rename_dict = {}
            for col in df.columns:
                if col in column_mappings:
                    rename_dict[col] = column_mappings[col]
            
            if rename_dict:
                df = df.rename(columns=rename_dict)
            
            # 값 매핑 적용
            for original_col, value_map in value_mappings.items():
                mapped_col = column_mappings.get(original_col, original_col)
                if mapped_col in df.columns:
                    df[mapped_col] = df[mapped_col].replace(value_map)
            
            merged_dfs.append(df)
        
        # 최종 통합
        if merged_dfs:
            merged_metadata = pd.concat(merged_dfs, ignore_index=True, sort=False)
            
            output_file = Path(output_dir) / 'merged_metadata.csv'
            merged_metadata.to_csv(output_file, index=False)
            self.log(f"Saved merged metadata to {output_file}")
            
            print(f"\n=== Merge Summary ===")
            print(f"Total rows: {len(merged_metadata)}")
            print(f"Total columns: {len(merged_metadata.columns)}")
            print(f"Data types: {merged_metadata['_data_type'].value_counts().to_dict()}")
            
            return merged_metadata
        
        return None

class AOIMatcher:
    """Count Matrix와 Metadata 매칭"""
    
    def __init__(self, verbose=True):
        self.verbose = verbose
        
    def log(self, message):
        if self.verbose:
            print(f"[INFO] {message}")
    
    def match_aoi_columns(self, matrix_csv, metadata_csv, output_dir):
        """
        사용자가 이미 처리한 count matrix CSV와 metadata CSV를 매칭
        
        matrix_csv: 첫 column은 gene, 나머지는 AOI columns
        metadata_csv: SegmentDisplayName column이 있는 metadata
        """
        # 파일 읽기
        self.log("Reading count matrix...")
        df_matrix = pd.read_csv(matrix_csv, index_col=0)  # 첫 column을 index로
        matrix_columns = df_matrix.columns.tolist()
        
        self.log("Reading metadata...")
        df_metadata = pd.read_csv(metadata_csv)
        
        if 'SegmentDisplayName' not in df_metadata.columns:
            raise ValueError("Metadata must have 'SegmentDisplayName' column")
        
        # 매칭 수행
        metadata_segments = df_metadata['SegmentDisplayName'].unique()
        
        matched = []
        unmatched_metadata = []
        fuzzy_matched = []
        
        # 매칭 딕셔너리 생성
        match_dict = {}
        
        for segment in metadata_segments:
            if pd.isna(segment):
                continue
                
            segment_str = str(segment)
            
            # 1. 정확 매칭
            if segment_str in matrix_columns:
                matched.append(segment_str)
                match_dict[segment_str] = segment_str
                
            else:
                # 2. Fuzzy 매칭 시도
                found = False
                
                # 파일명_AOI명 형식 찾기
                for col in matrix_columns:
                    # 여러 패턴 시도
                    patterns = [
                        col.endswith(f'_{segment_str}'),
                        col.endswith(f' | {segment_str}'),
                        segment_str in col.split('_')[-1] if '_' in col else False,
                        col.split(' | ')[-1] == segment_str if ' | ' in col else False,
                        # 추가 패턴
                        col.replace('_', ' | ').endswith(segment_str),
                        segment_str.replace(' | ', '_') in col,
                    ]
                    
                    if any(patterns):
                        fuzzy_matched.append((segment_str, col))
                        match_dict[segment_str] = col
                        found = True
                        break
                
                if not found:
                    unmatched_metadata.append(segment_str)
        
        # Matrix에만 있는 columns
        matched_matrix_cols = set(match_dict.values())
        unmatched_matrix = [col for col in matrix_columns if col not in matched_matrix_cols]
        
        # 결과 출력
        print("\n=== Matching Results ===")
        print(f"Exact matches: {len(matched)}")
        print(f"Fuzzy matches: {len(fuzzy_matched)}")
        print(f"Unmatched metadata: {len(unmatched_metadata)}")
        print(f"Unmatched matrix columns: {len(unmatched_matrix)}")
        
        # Fuzzy 매칭 상세
        if fuzzy_matched:
            print("\n=== Fuzzy Matches (Top 10) ===")
            for meta, matrix in fuzzy_matched[:10]:
                print(f"  {meta} → {matrix}")
        
        # 매칭된 metadata 저장
        df_metadata['matrix_column'] = df_metadata['SegmentDisplayName'].map(match_dict)
        df_matched = df_metadata[df_metadata['matrix_column'].notna()]
        
        output_path = Path(output_dir)
        output_path.mkdir(exist_ok=True)
        
        # 매칭된 metadata 저장
        matched_file = output_path / 'matched_metadata.csv'
        df_matched.to_csv(matched_file, index=False)
        self.log(f"Saved matched metadata to {matched_file}")
        
        # 매칭 리포트 저장
        report_file = output_path / 'matching_report.json'
        report = {
            'summary': {
                'total_metadata_segments': len(metadata_segments),
                'exact_matches': len(matched),
                'fuzzy_matches': len(fuzzy_matched),
                'unmatched_metadata': len(unmatched_metadata),
                'unmatched_matrix': len(unmatched_matrix)
            },
            'fuzzy_matches': fuzzy_matched,
            'unmatched_metadata': unmatched_metadata,
            'unmatched_matrix': unmatched_matrix[:50]  # 최대 50개만
        }
        
        with open(report_file, 'w') as f:
            json.dump(report, f, indent=2)
        self.log(f"Saved matching report to {report_file}")
        
        # 매칭된 matrix subset 저장 (옵션)
        if len(df_matched) > 0:
            matched_cols = df_matched['matrix_column'].unique()
            matched_cols = [col for col in matched_cols if col in matrix_columns]
            
            if matched_cols:
                df_matrix_matched = df_matrix[matched_cols]
                matrix_file = output_path / 'matched_count_matrix.csv'
                df_matrix_matched.to_csv(matrix_file)
                self.log(f"Saved matched count matrix to {matrix_file}")
        
        return df_matched, report

def main():
    parser = argparse.ArgumentParser(description='GeomX Data Processor v2')
    subparsers = parser.add_subparsers(dest='command', help='Commands')
    
    # analyze 명령
    analyze_parser = subparsers.add_parser('analyze', help='Analyze unique columns in xlsx files')
    analyze_parser.add_argument('input_dir', help='Directory containing xlsx files')
    analyze_parser.add_argument('-o', '--output', default='./output', help='Output directory')
    
    # merge 명령
    merge_parser = subparsers.add_parser('merge', help='Apply mappings and merge metadata')
    merge_parser.add_argument('input_dir', help='Directory containing xlsx files')
    merge_parser.add_argument('mapping_file', help='CSV file with column mappings')
    merge_parser.add_argument('-o', '--output', default='./output', help='Output directory')
    
    # match 명령
    match_parser = subparsers.add_parser('match', help='Match count matrix with metadata')
    match_parser.add_argument('matrix_csv', help='Count matrix CSV file')
    match_parser.add_argument('metadata_csv', help='Metadata CSV file')
    match_parser.add_argument('-o', '--output', default='./output', help='Output directory')
    
    args = parser.parse_args()
    
    if args.command == 'analyze':
        processor = GeomXProcessor()
        processor.process_directory(args.input_dir)
        processor.save_analysis_results(args.output)
        processor.generate_mapping_template(args.output)
        
    elif args.command == 'merge':
        processor = GeomXProcessor()
        processor.process_directory(args.input_dir)
        processor.apply_mappings_and_merge(args.mapping_file, args.output)
        
    elif args.command == 'match':
        matcher = AOIMatcher()
        matcher.match_aoi_columns(args.matrix_csv, args.metadata_csv, args.output)
        
    else:
        parser.print_help()

if __name__ == "__main__":
    main()